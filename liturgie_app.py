import calendar
import html
import io
import json
import random
import re
from copy import deepcopy
from concurrent.futures import ThreadPoolExecutor
from datetime import date, timedelta

import streamlit as st
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# -----------------------------------------------------------------------------
# Configuration
# -----------------------------------------------------------------------------
st.set_page_config(
    page_title="Programme liturgique",
    page_icon="⛪",
    layout="centered",
    initial_sidebar_state="collapsed",
)

APP_VERSION = "2026.08.24-vigiles-hybride-v2"

DEFAULT_FR = [f"F{i}" for i in range(1, 11)]
DEFAULT_MO = [f"M{i}" for i in range(1, 9)]
MONTHS = [
    "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre",
]
ROLE_LABELS = {
    None: "Libre (départ)",
    "LECTURE": "Lecture",
    "MONITION": "Monition + P.U.",
}

AELF_BASE_URL = "https://www.aelf.org"
AELF_CALENDAR = "romain"
AELF_TIMEOUT_SECONDS = 12
MAJOR_CELEBRATION_KEYWORDS = (
    # Solennités et grandes célébrations universelles / du Seigneur.
    # La détection finale reste fondée sur le libellé AELF du jour afin de
    # respecter les transferts de date et le calendrier liturgique réellement publié.
    "solennité", "pâques", "pentecôte", "ascension", "tous les saints",
    "christ roi", "nativité du seigneur", "sainte famille", "épiphanie",
    "trinité", "saint-sacrement", "assomption", "immaculée conception",
    "conception immaculée", "sainte marie, mère de dieu", "sainte marie mère de dieu",
    "annonciation", "saint joseph", "sacré-cœur", "sacré coeur",
    "saint pierre et saint paul", "saints pierre et paul",
    "nativité de saint jean baptiste", "présentation du seigneur",
    "exaltation de la sainte croix", "jeudi saint", "vendredi saint",
    "vigile pascale", "veillée pascale",
)

# Interface mobile-first
st.markdown(
    """
    <style>
      .block-container {max-width: 900px; padding-top: 1rem; padding-bottom: 5rem;}
      div.stButton > button, div.stDownloadButton > button {
        width: 100%; min-height: 3rem; border-radius: 12px; font-weight: 700;
      }
      [data-testid="stMetricValue"] {font-size: 1.25rem;}
      .lit-card {border: 1px solid rgba(128,128,128,.25); border-radius: 14px;
                 padding: 14px; margin: 10px 0;}
      .lit-date {font-size: 1.15rem; font-weight: 800; margin-bottom: 8px;}
      .lit-grid {display:grid; grid-template-columns:1fr; gap:10px;}
      .lit-section {padding:10px; border-radius:10px; background:rgba(128,128,128,.07);}
      .lit-label {font-weight:800; margin-bottom:5px;}
      .small-note {font-size:.9rem; opacity:.8;}
      @media (min-width: 760px) {
        .lit-grid {grid-template-columns:1.15fr 1fr 1fr 1fr;}
      }
    </style>
    """,
    unsafe_allow_html=True,
)

# -----------------------------------------------------------------------------
# Etat et utilitaires
# -----------------------------------------------------------------------------
def blank_person():
    return {
        "next_role": None,
        "last_reading": None,
        "last_service": None,
        "last_announcement": None,
        "reading_count": 0,
        "monition_count": 0,
        "announcement_count": 0,
    }


def code_lang(code):
    if isinstance(code, str) and code.startswith("F"):
        return "FR"
    if isinstance(code, str) and code.startswith("M"):
        return "MO"
    return None


def code_number(code):
    try:
        return int(code[1:])
    except (TypeError, ValueError):
        return 10**9


def sort_codes(codes):
    return sorted(dict.fromkeys(codes), key=lambda c: (code_lang(c) or "ZZ", code_number(c), c))


def clean_name(value):
    return " ".join(str(value or "").strip().split())


def member_name(state, code):
    return clean_name(state.get("names", {}).get(code, ""))


def display_name(state, code):
    name = member_name(state, code)
    return name if name else "Nom à renseigner"


def all_member_codes(state):
    return sort_codes(list(state.get("people", {}).keys()))


def missing_active_names(state):
    return [
        code for lang in ("FR", "MO") for code in active_codes(state, lang)
        if not member_name(state, code)
    ]


def validate_name(state, name, exclude_code=None):
    name = clean_name(name)
    if not name:
        raise RuntimeError("Le nom du membre ne peut pas être vide.")
    wanted = name.casefold()
    for code in all_member_codes(state):
        if code == exclude_code:
            continue
        current = member_name(state, code)
        if current and current.casefold() == wanted:
            raise RuntimeError(
                "Ce nom est déjà utilisé. Ajoutez une précision (par exemple n°1 / n°2) "
                "pour distinguer les deux personnes."
            )
    return name


def set_member_name(state, code, name):
    if code not in state.get("people", {}):
        raise RuntimeError("Membre introuvable.")
    state.setdefault("names", {})[code] = validate_name(state, name, exclude_code=code)


def member_format(state):
    return lambda code: display_name(state, code) if code != "Auto" else "Auto"


def display_row(row, state):
    shown = dict(row)
    pattern = re.compile(r"\b[FM]\d+\b")
    for field in ["Lecteurs", "Monition introductive + P.U.", "Chargés d’annonce"]:
        text = str(shown.get(field, ""))
        shown[field] = pattern.sub(lambda m: display_name(state, m.group(0)), text)
    return shown


def initial_state():
    return {
        "version": 4,
        "members": {"FR": DEFAULT_FR[:], "MO": DEFAULT_MO[:]},
        "inactive_members": {"FR": [], "MO": []},
        "names": {code: "" for code in DEFAULT_FR + DEFAULT_MO},
        "people": {code: blank_person() for code in DEFAULT_FR + DEFAULT_MO},
        "reading_cycle_seen": {"FR": [], "MO": []},
        "reading_pairs": [],
        "monition_pairs": [],
        "next_first_language": "FR",
        "history": [],
    }


def active_codes(state, lang):
    return state.get("members", {}).get(lang, [])[:]


def normalize_state(raw):
    state = initial_state()
    if not isinstance(raw, dict):
        return state

    # Champs généraux conservés lors d'un import d'une ancienne version.
    for key in ["reading_cycle_seen", "reading_pairs", "monition_pairs",
                "next_first_language", "history"]:
        if key in raw:
            state[key] = deepcopy(raw[key])

    raw_people = raw.get("people", {}) if isinstance(raw.get("people", {}), dict) else {}

    # Version 3 : la liste des membres actifs est stockée dans l'état.
    raw_members = raw.get("members")
    if isinstance(raw_members, dict):
        fr_members = [c for c in raw_members.get("FR", []) if code_lang(c) == "FR"]
        mo_members = [c for c in raw_members.get("MO", []) if code_lang(c) == "MO"]
    else:
        # Compatibilité avec les états des versions précédentes.
        inferred_fr = [c for c in raw_people if code_lang(c) == "FR"]
        inferred_mo = [c for c in raw_people if code_lang(c) == "MO"]
        fr_members = inferred_fr or DEFAULT_FR[:]
        mo_members = inferred_mo or DEFAULT_MO[:]

    state["members"] = {"FR": sort_codes(fr_members), "MO": sort_codes(mo_members)}

    raw_inactive = raw.get("inactive_members", {})
    if isinstance(raw_inactive, dict):
        state["inactive_members"] = {
            "FR": sort_codes([c for c in raw_inactive.get("FR", []) if code_lang(c) == "FR" and c not in state["members"]["FR"]]),
            "MO": sort_codes([c for c in raw_inactive.get("MO", []) if code_lang(c) == "MO" and c not in state["members"]["MO"]]),
        }
    else:
        state["inactive_members"] = {"FR": [], "MO": []}

    all_codes = set(DEFAULT_FR + DEFAULT_MO)
    all_codes.update(raw_people.keys())
    all_codes.update(state["members"]["FR"] + state["members"]["MO"])
    all_codes.update(state["inactive_members"]["FR"] + state["inactive_members"]["MO"])
    state["people"] = {code: blank_person() for code in sort_codes(all_codes) if code_lang(code)}
    for code, pdata in raw_people.items():
        if code in state["people"] and isinstance(pdata, dict):
            state["people"][code].update(pdata)

    raw_names = raw.get("names", {}) if isinstance(raw.get("names", {}), dict) else {}
    state["names"] = {code: clean_name(raw_names.get(code, "")) for code in state["people"]}

    # Un membre supprimé ne doit plus bloquer le cycle de lecture actif.
    for lang in ["FR", "MO"]:
        seen = state.get("reading_cycle_seen", {}).get(lang, [])
        state["reading_cycle_seen"][lang] = [c for c in seen if c in state["members"][lang]]

    state["version"] = 4
    return state


def next_member_code(state, lang):
    prefix = "F" if lang == "FR" else "M"
    nums = [code_number(c) for c in state.get("people", {}) if code_lang(c) == lang]
    n = max([x for x in nums if x < 10**9], default=0) + 1
    return f"{prefix}{n}"


def add_member(state, lang, name):
    name = validate_name(state, name)
    code = next_member_code(state, lang)
    state["people"][code] = blank_person()
    state.setdefault("names", {})[code] = name
    state["members"][lang].append(code)
    state["members"][lang] = sort_codes(state["members"][lang])
    return code


def remove_member(state, code):
    lang = code_lang(code)
    if lang not in ("FR", "MO") or code not in state["members"][lang]:
        raise RuntimeError("Membre actif introuvable.")
    if len(state["members"][lang]) <= 3:
        raise RuntimeError(
            "Il faut conserver au moins 3 membres actifs dans chaque catégorie "
            "pour assurer Lecture, Monition/P.U. et Annonce sans cumul le même dimanche."
        )
    state["members"][lang].remove(code)
    if code not in state["inactive_members"][lang]:
        state["inactive_members"][lang].append(code)
        state["inactive_members"][lang] = sort_codes(state["inactive_members"][lang])
    state["reading_cycle_seen"][lang] = [c for c in state["reading_cycle_seen"][lang] if c != code]


def reactivate_member(state, code):
    lang = code_lang(code)
    if lang not in ("FR", "MO") or code not in state["inactive_members"][lang]:
        raise RuntimeError("Membre retiré introuvable.")
    if not member_name(state, code):
        raise RuntimeError("Renseignez d’abord le nom de ce membre avant de le réactiver.")
    state["inactive_members"][lang].remove(code)
    if code not in state["members"][lang]:
        state["members"][lang].append(code)
        state["members"][lang] = sort_codes(state["members"][lang])
    state["people"].setdefault(code, blank_person())


def apply_bulk_names(state, fr_names, mo_names):
    groups = {"FR": fr_names, "MO": mo_names}
    proposed = {}
    for lang, names in groups.items():
        codes = active_codes(state, lang)
        cleaned = [clean_name(x) for x in names]
        if len(cleaned) != len(codes):
            label = "francophones" if lang == "FR" else "mooréphones"
            raise RuntimeError(f"Il faut saisir exactement {len(codes)} noms {label}.")
        if any(not x for x in cleaned):
            raise RuntimeError("Aucun nom ne doit être vide.")
        proposed.update(dict(zip(codes, cleaned)))

    all_new = [x.casefold() for x in proposed.values()]
    if len(all_new) != len(set(all_new)):
        raise RuntimeError("Deux membres actifs ont le même nom. Ajoutez une précision pour les distinguer.")

    inactive_names = {
        member_name(state, code).casefold()
        for lang in ("FR", "MO")
        for code in state.get("inactive_members", {}).get(lang, [])
        if member_name(state, code)
    }
    collision = inactive_names.intersection(all_new)
    if collision:
        raise RuntimeError(
            "Un nom saisi appartient déjà à un membre retiré. Réactivez ce membre ou utilisez un nom distinct."
        )

    state.setdefault("names", {}).update(proposed)


def reset_rotation_keep_members(state):
    state = normalize_state(deepcopy(state))
    state["people"] = {code: blank_person() for code in state["people"]}
    state["reading_cycle_seen"] = {"FR": [], "MO": []}
    state["reading_pairs"] = []
    state["monition_pairs"] = []
    state["next_first_language"] = "FR"
    state["history"] = []
    return state


def as_date(value):
    return date.fromisoformat(value) if value else None


def days_since(value, today):
    return 10000 if not value else (today - as_date(value)).days


def sundays(year, month):
    cal = calendar.Calendar(firstweekday=calendar.MONDAY)
    return [
        d for d in cal.itermonthdates(year, month)
        if d.month == month and d.weekday() == 6
    ]


def month_dates(year, month):
    """Toutes les dates civiles du mois, dans l'ordre."""
    last_day = calendar.monthrange(year, month)[1]
    return [date(year, month, day) for day in range(1, last_day + 1)]


def _compact_spaces(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def _extract_reference_from_heading(text):
    """Extrait la référence biblique finale d'un titre AELF.

    Exemple : « ... » (Ez 33, 7-9) -> Ez 33, 7-9
    """
    text = _compact_spaces(text)
    match = re.search(r"\(([^()]*)\)\s*$", text)
    if match:
        return _compact_spaces(match.group(1))
    return ""


def _extract_aelf_celebration(soup):
    """Essaie de récupérer le nom de la célébration affichée par l'AELF."""
    lines = [_compact_spaces(x) for x in soup.stripped_strings]
    try:
        end = lines.index("Lectures de la messe")
    except ValueError:
        end = min(len(lines), 100)
    candidates = lines[:end]

    # Le libellé principal AELF contient très souvent « Année A/B/C ».
    for value in reversed(candidates):
        if "Année " in value and len(value) < 180:
            return value

    # Repli pour certaines solennités/fêtes dont le libellé est séparé.
    wanted = (
        "dimanche", "solennité", "nativité", "tous les saints", "christ roi",
        "sainte famille", "épiphanie", "pentecôte", "pâques", "ascension",
        "assomption", "immaculée conception", "conception immaculée", "trinité", "saint-sacrement",
        "sainte marie, mère de dieu", "sainte marie mère de dieu",
        "annonciation", "saint joseph", "sacré-cœur", "sacré coeur",
        "saint pierre", "saint paul", "saint jean baptiste",
        "présentation du seigneur", "exaltation de la sainte croix",
        "jeudi saint", "vendredi saint", "vigile pascale", "veillée pascale",
    )
    for value in reversed(candidates):
        low = value.casefold()
        if any(word in low for word in wanted) and len(value) < 180:
            return value
    return ""


def _is_major_celebration(label):
    low = _compact_spaces(label).casefold()
    return any(keyword in low for keyword in MAJOR_CELEBRATION_KEYWORDS)


@st.cache_data(ttl=6 * 3600, show_spinner=False)
def fetch_aelf_mass(date_iso, calendar_code=AELF_CALENDAR):
    """Charge les références de la messe du jour depuis l'AELF.

    La fonction ne récupère que les références, pas le texte intégral des lectures.
    Elle renvoie toujours un dictionnaire afin que l'utilisateur puisse corriger
    manuellement si le site n'est pas accessible ou si une variante locale s'applique.
    """
    url = f"{AELF_BASE_URL}/{date_iso}/{calendar_code}/messe"
    result = {
        "r1": "", "r2": "", "ev": "", "celebration": "",
        "source_url": url, "status": "error", "error": "",
    }
    try:
        response = requests.get(
            url,
            timeout=AELF_TIMEOUT_SECONDS,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (Linux; Android 13) AppleWebKit/537.36 "
                    "Chrome/124 Safari/537.36 ProgrammeLiturgique/1.0"
                )
            },
        )
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")
        result["celebration"] = _extract_aelf_celebration(soup)

        wanted = {
            "Première lecture": "r1",
            "Deuxième lecture": "r2",
            "Évangile": "ev",
        }
        for heading in soup.find_all(re.compile(r"^h[1-6]$")):
            title = _compact_spaces(heading.get_text(" ", strip=True))
            if title not in wanted:
                continue
            key = wanted[title]
            # Sur l'AELF, le titre suivant contient la référence entre parenthèses.
            next_heading = heading.find_next(re.compile(r"^h[1-6]$"))
            if next_heading is not None:
                ref = _extract_reference_from_heading(next_heading.get_text(" ", strip=True))
                if ref and not result[key]:
                    result[key] = ref

        # Repli : les liens du sommaire contiennent aussi les références.
        if not (result["r1"] and result["r2"] and result["ev"]):
            for item in soup.find_all(["a", "li"]):
                line = _compact_spaces(item.get_text(" ", strip=True))
                for label, key in wanted.items():
                    if result[key] or not line.startswith(label):
                        continue
                    # La référence est généralement à la fin après le titre entre guillemets.
                    tail = re.sub(r"^" + re.escape(label), "", line).strip()
                    tail = re.sub(r"^«.*?»\s*", "", tail).strip()
                    if tail:
                        result[key] = tail

        if result["r1"] and result["r2"] and result["ev"]:
            result["status"] = "ok"
        else:
            missing = [
                label for label, key in (("1re lecture", "r1"), ("2e lecture", "r2"), ("Évangile", "ev"))
                if not result[key]
            ]
            result["status"] = "partial"
            result["error"] = "Référence(s) non détectée(s) : " + ", ".join(missing)
    except Exception as exc:
        result["error"] = f"AELF inaccessible : {exc}"
    return result


def refs_to_text(refs, dates):
    lines = []
    for d in dates:
        item = refs.get(d.isoformat(), {})
        lines.append(
            f"{d.isoformat()} | {item.get('r1', '')} | {item.get('r2', '')} | {item.get('ev', '')}"
        )
    return "\n".join(lines)


def load_automatic_refs(dates, calendar_code=AELF_CALENDAR):
    refs = {}
    info = []
    dates = list(dates)
    if not dates:
        return refs, info

    # Un mois complet peut contenir 31 dates. La mise en cache AELF évite de
    # recharger inutilement les mêmes pages lors des reruns Streamlit.
    workers = min(8, len(dates))
    with ThreadPoolExecutor(max_workers=workers) as pool:
        items = list(pool.map(lambda d: fetch_aelf_mass(d.isoformat(), calendar_code), dates))

    for d, item in zip(dates, items):
        refs[d.isoformat()] = {
            "r1": item.get("r1", ""),
            "r2": item.get("r2", ""),
            "ev": item.get("ev", ""),
        }
        info.append({
            "date": d.isoformat(),
            "celebration": item.get("celebration", ""),
            "status": item.get("status", "error"),
            "error": item.get("error", ""),
            "source_url": item.get("source_url", ""),
            "major": _is_major_celebration(item.get("celebration", "")),
        })
    return refs, info


def discover_month_liturgies(year, month, include_weekday_major=True, calendar_code=AELF_CALENDAR):
    """Construit les dates à programmer pour un mois.

    Tous les dimanches sont toujours inclus. Si ``include_weekday_major`` est
    actif, l'application analyse aussi les autres jours du mois et ajoute les
    grandes célébrations détectées par l'AELF. Cela permet notamment de suivre
    automatiquement les transferts liturgiques publiés par l'AELF.
    """
    all_days = month_dates(year, month)
    all_refs, all_info = load_automatic_refs(all_days, calendar_code)
    info_by_date = {item["date"]: item for item in all_info}

    selected = []
    for d in all_days:
        item = info_by_date.get(d.isoformat(), {})
        if d.weekday() == 6:
            selected.append(d)
        elif include_weekday_major and item.get("major"):
            selected.append(d)

    # Dédupliquer et garantir l'ordre chronologique.
    selected = sorted(set(selected))
    refs = {d.isoformat(): all_refs.get(d.isoformat(), {"r1": "", "r2": "", "ev": ""}) for d in selected}
    info = [info_by_date.get(d.isoformat(), {"date": d.isoformat(), "celebration": "", "major": False, "status": "error", "error": "", "source_url": ""}) for d in selected]
    return selected, refs, info



# -----------------------------------------------------------------------------
# Calendrier hybride : dimanches automatiques + célébrations supplémentaires
# manuelles, avec références AELF automatiques (y compris les vigiles).
# -----------------------------------------------------------------------------
READING_HEADINGS = {"Première lecture", "Deuxième lecture", "Épître", "Évangile", "Lecture"}


def _slug(value):
    value = _compact_spaces(value).casefold()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-") or "messe"


def _clean_celebration_label(value):
    value = _compact_spaces(value)
    value = re.sub(r"\s+—\s+Année\s+[ABC].*$", "", value, flags=re.I)
    return value.strip(" —")


@st.cache_data(ttl=6 * 3600, show_spinner=False)
def fetch_aelf_forms(date_iso, calendar_code=AELF_CALENDAR):
    """Récupère les différents formulaires de messe d'une date AELF.

    Contrairement à ``fetch_aelf_mass``, cette fonction conserve les sections
    distinctes (messe de la veille, de la nuit, du jour, veillée pascale...) et
    toutes les références de lecture utiles à la programmation manuelle.
    """
    url = f"{AELF_BASE_URL}/{date_iso}/{calendar_code}/messe"
    result = {
        "date": date_iso,
        "celebration": "",
        "forms": [],
        "source_url": url,
        "status": "error",
        "error": "",
    }
    try:
        response = requests.get(
            url,
            timeout=AELF_TIMEOUT_SECONDS,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (Linux; Android 13) AppleWebKit/537.36 "
                    "Chrome/124 Safari/537.36 ProgrammeLiturgique/2.0"
                )
            },
        )
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")
        result["celebration"] = _extract_aelf_celebration(soup)

        forms = []
        current = None
        headings = soup.find_all(re.compile(r"^h[1-6]$"))
        for heading in headings:
            text = _compact_spaces(heading.get_text(" ", strip=True))
            low = text.casefold()
            if heading.name == "h1":
                is_form = (
                    low == "lectures de la messe"
                    or "messe" in low
                    or "veillée pascale" in low
                    or "veillee pascale" in low
                    or "vigile pascale" in low
                )
                if is_form:
                    current = {"title": text, "readings": []}
                    forms.append(current)
                else:
                    current = None
                continue

            if current is None or text not in READING_HEADINGS:
                continue

            nxt = heading.find_next(re.compile(r"^h[1-6]$"))
            if nxt is None or nxt.name == "h1":
                continue
            ref = _extract_reference_from_heading(nxt.get_text(" ", strip=True))
            if not ref:
                continue
            item = {"label": text, "ref": ref}
            if item not in current["readings"]:
                current["readings"].append(item)

        # Supprimer les formulaires vides et dédoublonner les sections répétées.
        clean_forms = []
        seen = set()
        for form in forms:
            if not form["readings"]:
                continue
            signature = (form["title"].casefold(), tuple((r["label"], r["ref"]) for r in form["readings"]))
            if signature in seen:
                continue
            seen.add(signature)
            clean_forms.append(form)
        result["forms"] = clean_forms
        if clean_forms:
            result["status"] = "ok"
        else:
            result["status"] = "partial"
            result["error"] = "Aucun formulaire de messe détaillé n'a été détecté."
    except Exception as exc:
        result["error"] = f"AELF inaccessible : {exc}"
    return result


def _form_low(form):
    return _compact_spaces(form.get("title", "")).casefold()


def _is_vigil_form(form):
    low = _form_low(form)
    return "messe de la veille au soir" in low or "veillée pascale" in low or "veillee pascale" in low or "vigile pascale" in low


def _is_paschal_vigil(form):
    low = _form_low(form)
    return "veillée pascale" in low or "veillee pascale" in low or "vigile pascale" in low


def _is_christmas_night(form, celebration):
    return "nativité du seigneur" in _compact_spaces(celebration).casefold() and "messe de la nuit" in _form_low(form)


def _preferred_day_form(page):
    forms = page.get("forms", [])
    for wanted in ("messe du jour de pâques", "messe du jour", "lectures de la messe"):
        for form in forms:
            if wanted in _form_low(form):
                return form
    for form in forms:
        low = _form_low(form)
        if not _is_vigil_form(form) and "messe de la nuit" not in low and "messe de l'aurore" not in low:
            return form
    return forms[0] if forms else None


def _form_standard_refs(form):
    readings = form.get("readings", []) if form else []
    firsts = [r for r in readings if r["label"] == "Première lecture"]
    seconds = [r for r in readings if r["label"] in ("Deuxième lecture", "Épître")]
    gospels = [r for r in readings if r["label"] == "Évangile"]
    other = [r for r in readings if r["label"] == "Lecture"]
    r1 = (firsts or other or readings[:1])
    r2 = seconds
    return {
        "r1": r1[0]["ref"] if r1 else "",
        "r2": r2[0]["ref"] if r2 else "",
        "ev": gospels[0]["ref"] if gospels else "",
    }


def _form_reference_lines(form, special="standard"):
    if not form:
        return []
    readings = form.get("readings", [])
    if special == "paschal_vigil":
        lines = []
        num = 0
        for item in readings:
            if item["label"] == "Évangile":
                continue
            if item["label"] == "Épître":
                label = "Épître"
            else:
                num += 1
                label = f"Lecture {num}"
            lines.append({"label": label, "ref": item["ref"], "kind": "reading", "source_label": item["label"]})
        for item in readings:
            if item["label"] == "Évangile":
                lines.append({"label": "Év.", "ref": item["ref"], "kind": "gospel", "source_label": item["label"]})
                break
        return lines

    refs = _form_standard_refs(form)
    lines = []
    if refs["r1"]:
        lines.append({"label": "1re", "ref": refs["r1"], "kind": "reading", "source_label": "Première lecture"})
    if refs["r2"]:
        lines.append({"label": "2e", "ref": refs["r2"], "kind": "reading", "source_label": "Deuxième lecture"})
    if refs["ev"]:
        lines.append({"label": "Év.", "ref": refs["ev"], "kind": "gospel", "source_label": "Évangile"})
    return lines


def _make_service(service_date, source_date, page, form, kind, automatic_people, order, default_include=True):
    celebration = _clean_celebration_label(page.get("celebration", "")) or "Célébration"
    form_title = _compact_spaces((form or {}).get("title", "Lectures de la messe"))
    special = "paschal_vigil" if _is_paschal_vigil(form or {}) else "standard"
    source_url = page.get("source_url", f"{AELF_BASE_URL}/{source_date.isoformat()}/{AELF_CALENDAR}/messe")
    sid = f"{service_date.isoformat()}_{kind}_{_slug(form_title)}"
    return {
        "id": sid,
        "date": service_date.isoformat(),
        "source_date": source_date.isoformat(),
        "kind": kind,
        "automatic_people": bool(automatic_people),
        "order": int(order),
        "default_include": bool(default_include),
        "celebration": celebration,
        "form_title": form_title,
        "form": deepcopy(form or {"title": form_title, "readings": []}),
        "special": special,
        "source_url": source_url,
        "status": page.get("status", "error"),
        "error": page.get("error", ""),
    }


@st.cache_data(ttl=6 * 3600, show_spinner=False)
def discover_hybrid_services(year, month, calendar_code=AELF_CALENDAR):
    """Détecte les services du mois selon les règles de l'application.

    - Dimanches : références + affectations automatiques.
    - Grandes célébrations en semaine : références automatiques, personnes manuelles.
    - Vigiles proposées par l'AELF : ajoutées la veille, personnes manuelles.
    - Noël : veille au soir et messe de la nuit ajoutées le 24 décembre ;
      aurore et jour restent disponibles le 25 décembre.
    - Pâques : la Veillée pascale est ajoutée le samedi, tandis que la messe du
      jour de Pâques reste la célébration dominicale automatique.
    """
    days = month_dates(year, month)
    if not days:
        return []
    fetch_days = days + [days[-1] + timedelta(days=1)]
    workers = min(8, len(fetch_days))
    with ThreadPoolExecutor(max_workers=workers) as pool:
        pages = list(pool.map(lambda d: fetch_aelf_forms(d.isoformat(), calendar_code), fetch_days))
    page_by_date = {d.isoformat(): p for d, p in zip(fetch_days, pages)}

    services = []
    seen_ids = set()

    def add(service):
        if service and service["id"] not in seen_ids:
            seen_ids.add(service["id"])
            services.append(service)

    for d in days:
        page = page_by_date.get(d.isoformat(), {})
        celebration_low = _compact_spaces(page.get("celebration", "")).casefold()
        day_form = _preferred_day_form(page)

        # 1) Messe dominicale : personnes automatiques.
        if d.weekday() == 6 and day_form:
            add(_make_service(d, d, page, day_form, "sunday", True, 20, True))

        # 2) Grande célébration en semaine : personnes manuelles.
        if d.weekday() != 6 and _is_major_celebration(page.get("celebration", "")) and day_form:
            add(_make_service(d, d, page, day_form, "weekday_major", False, 20, True))

        # Noël comporte plusieurs formulaires. L'aurore est proposée séparément.
        if "nativité du seigneur" in celebration_low:
            for form in page.get("forms", []):
                low = _form_low(form)
                if "messe de l'aurore" in low:
                    add(_make_service(d, d, page, form, "christmas_dawn", False, 10, False))

        # 3) Vigiles célébrées ce soir : elles se trouvent sur la page AELF du lendemain.
        next_date = d + timedelta(days=1)
        next_page = page_by_date.get(next_date.isoformat())
        if next_page:
            next_celebration = next_page.get("celebration", "")
            for form in next_page.get("forms", []):
                if _is_paschal_vigil(form):
                    add(_make_service(d, next_date, next_page, form, "paschal_vigil", False, 80, True))
                elif "messe de la veille au soir" in _form_low(form):
                    add(_make_service(d, next_date, next_page, form, "vigil", False, 80, True))
                elif _is_christmas_night(form, next_celebration):
                    add(_make_service(d, next_date, next_page, form, "christmas_night", False, 90, True))

    services.sort(key=lambda s: (s["date"], s["order"], s["id"]))
    return services


def service_reference_text(service):
    lines = _form_reference_lines(service.get("form"), service.get("special", "standard"))
    return "\n".join(f"{x['label']} : {x['ref']}" for x in lines)


def service_display_title(service):
    title = service.get("celebration", "Célébration")
    form = service.get("form_title", "")
    if form and form.casefold() not in ("lectures de la messe", "messe du jour"):
        return f"{title} — {form}"
    return title


def _standard_manual_slots(service, first_choice=0):
    readings = service.get("form", {}).get("readings", [])
    firsts = [r for r in readings if r["label"] == "Première lecture"]
    seconds = [r for r in readings if r["label"] in ("Deuxième lecture", "Épître")]
    others = [r for r in readings if r["label"] == "Lecture"]
    slots = []
    if firsts:
        idx = min(max(int(first_choice or 0), 0), len(firsts) - 1)
        slots.append({"label": "1re", "ref": firsts[idx]["ref"]})
    elif others:
        slots.append({"label": "1re", "ref": others[0]["ref"]})
    if seconds:
        slots.append({"label": "2e", "ref": seconds[0]["ref"]})
    elif len(others) > 1:
        slots.append({"label": "2e", "ref": others[1]["ref"]})
    return slots, firsts


def _service_key(service, suffix):
    return f"svc_{_slug(service['id'])}_{suffix}"


def collect_manual_service_payload(service, state):
    """Affiche les contrôles manuels d'une célébration supplémentaire et renvoie les choix."""
    payload = {"readings": []}
    all_active = active_codes(state, "FR") + active_codes(state, "MO")

    st.caption(service_reference_text(service) or "Références non disponibles — utilisez Recharger AELF.")
    if service.get("source_url"):
        st.caption(f"Source : {service['source_url']}")

    if service.get("special") == "paschal_vigil":
        st.info(
            "Veillée pascale : toutes les lectures bibliques détectées sont affichées. "
            "Choisissez manuellement la langue et le lecteur de chaque lecture proclamée. "
            "L'Évangile n'est pas attribué à un lecteur du groupe."
        )
        refs = [x for x in _form_reference_lines(service.get("form"), "paschal_vigil") if x["kind"] == "reading"]
        for idx, item in enumerate(refs):
            mandatory = item["label"] == "Épître" or item["ref"].replace(" ", "").startswith("Ex14")
            include = st.checkbox(
                f"{item['label']} — {item['ref']}",
                value=True,
                disabled=mandatory,
                key=_service_key(service, f"read_include_{idx}"),
            )
            if not include:
                continue
            default_lang = "FR" if idx % 2 == 0 else "MO"
            lang = st.selectbox(
                f"Langue — {item['label']}", ["FR", "MO"],
                index=0 if default_lang == "FR" else 1,
                key=_service_key(service, f"read_lang_{idx}"),
            )
            code = st.selectbox(
                f"Lecteur — {item['label']}", active_codes(state, lang),
                index=None, placeholder="Choisir un nom",
                format_func=member_format(state),
                key=_service_key(service, f"read_code_{idx}_{lang}"),
            )
            payload["readings"].append({"label": item["label"], "ref": item["ref"], "lang": lang, "code": code})
    else:
        readings = service.get("form", {}).get("readings", [])
        firsts = [r for r in readings if r["label"] == "Première lecture"]
        first_choice = 0
        if len(firsts) > 1:
            options = list(range(len(firsts)))
            first_choice = st.selectbox(
                "Choix de la 1re lecture", options,
                format_func=lambda i: firsts[i]["ref"],
                key=_service_key(service, "first_choice"),
            )
        slots, _ = _standard_manual_slots(service, first_choice)
        if not slots:
            st.warning("Aucune lecture avant l'Évangile n'a été détectée pour ce formulaire.")
        first_lang = st.selectbox(
            "Langue de la 1re lecture", ["FR", "MO"],
            key=_service_key(service, "first_lang"),
        )
        for idx, item in enumerate(slots):
            lang = first_lang if idx == 0 else ("MO" if first_lang == "FR" else "FR")
            code = st.selectbox(
                f"{item['label']} lecture — {lang} — {item['ref']}",
                active_codes(state, lang), index=None, placeholder="Choisir un nom",
                format_func=member_format(state),
                key=_service_key(service, f"standard_read_{idx}_{lang}"),
            )
            payload["readings"].append({"label": item["label"], "ref": item["ref"], "lang": lang, "code": code})

    st.markdown("**Monition introductive + P.U.**")
    c1, c2 = st.columns(2)
    with c1:
        payload["monition_fr"] = st.selectbox(
            "FR — Monition/P.U.", active_codes(state, "FR"), index=None, placeholder="Choisir un nom",
            format_func=member_format(state), key=_service_key(service, "monition_fr")
        )
    with c2:
        payload["monition_mo"] = st.selectbox(
            "MO — Monition/P.U.", active_codes(state, "MO"), index=None, placeholder="Choisir un nom",
            format_func=member_format(state), key=_service_key(service, "monition_mo")
        )

    st.markdown("**Annonces**")
    c1, c2 = st.columns(2)
    with c1:
        payload["annonce_fr"] = st.selectbox(
            "FR — Annonce", active_codes(state, "FR"), index=None, placeholder="Choisir un nom",
            format_func=member_format(state), key=_service_key(service, "annonce_fr")
        )
    with c2:
        payload["annonce_mo"] = st.selectbox(
            "MO — Annonce", active_codes(state, "MO"), index=None, placeholder="Choisir un nom",
            format_func=member_format(state), key=_service_key(service, "annonce_mo")
        )
    return payload


def _validate_active_code(state, code, lang=None):
    if not code:
        raise RuntimeError("Une affectation manuelle est incomplète.")
    actual = code_lang(code)
    if actual not in ("FR", "MO") or code not in active_codes(state, actual):
        raise RuntimeError(f"{display_name(state, code)} n'est pas un membre actif.")
    if lang and actual != lang:
        raise RuntimeError(f"{display_name(state, code)} n'appartient pas à la catégorie {lang}.")


def _apply_manual_service(state, service, payload):
    today = date.fromisoformat(service["date"])
    readings = payload.get("readings", [])
    if not readings:
        raise RuntimeError(f"{service_display_title(service)} : aucun lecteur biblique n'a été choisi.")

    required = [payload.get("monition_fr"), payload.get("monition_mo"), payload.get("annonce_fr"), payload.get("annonce_mo")]
    if any(x is None for x in required) or any(x.get("code") is None for x in readings):
        raise RuntimeError(f"{service_display_title(service)} : complétez toutes les affectations manuelles.")

    used = [x["code"] for x in readings] + required
    if len(used) != len(set(used)):
        raise RuntimeError(f"{service_display_title(service)} : une même personne ne peut pas cumuler deux fonctions à la même célébration.")

    # Vérifier l'alternance individuelle Lecture ↔ Monition/P.U.
    for item in readings:
        code = item["code"]
        _validate_active_code(state, code, item.get("lang"))
        nxt = state["people"][code]["next_role"]
        if nxt not in (None, "LECTURE"):
            raise RuntimeError(
                f"{service_display_title(service)} : {display_name(state, code)} doit faire Monition/P.U. à son prochain passage, pas une lecture."
            )
    for lang, key in (("FR", "monition_fr"), ("MO", "monition_mo")):
        code = payload[key]
        _validate_active_code(state, code, lang)
        nxt = state["people"][code]["next_role"]
        if nxt not in (None, "MONITION"):
            raise RuntimeError(
                f"{service_display_title(service)} : {display_name(state, code)} doit faire une lecture à son prochain passage, pas Monition/P.U."
            )
    _validate_active_code(state, payload["annonce_fr"], "FR")
    _validate_active_code(state, payload["annonce_mo"], "MO")

    # Appliquer la rotation. Les jours supplémentaires ne modifient PAS
    # l'alternance FR/MO des dimanches, mais leurs passages individuels comptent.
    for item in readings:
        assign(state, item["code"], "LECTURE", today)
    fr_reads = [x["code"] for x in readings if x.get("lang") == "FR"]
    mo_reads = [x["code"] for x in readings if x.get("lang") == "MO"]
    if len(fr_reads) == 1 and len(mo_reads) == 1:
        state["reading_pairs"].append([fr_reads[0], mo_reads[0]])

    assign(state, payload["monition_fr"], "MONITION", today)
    assign(state, payload["monition_mo"], "MONITION", today)
    state["monition_pairs"].append([payload["monition_fr"], payload["monition_mo"]])
    assign(state, payload["annonce_fr"], "ANNONCE", today)
    assign(state, payload["annonce_mo"], "ANNONCE", today)

    ref_lines = [f"⭐ {service_display_title(service)}"]
    # Afficher exactement les lectures retenues manuellement, puis l'Évangile.
    ref_lines.extend(f"{x['label']} : {x['ref']}" for x in readings)
    gospel = next((x for x in service.get("form", {}).get("readings", []) if x.get("label") == "Évangile"), None)
    if gospel:
        ref_lines.append(f"Év. : {gospel['ref']}")
    reader_lines = [f"{x['label']} : {x['lang']} — {x['code']}" for x in readings]
    return {
        "date": service["date"],
        "service_id": service["id"],
        "Dx": f"J{today.day}",
        "Réf.D": "\n".join(ref_lines),
        "Lecteurs": "\n".join(reader_lines),
        "Monition introductive + P.U.": f"FR — {payload['monition_fr']}\nMO — {payload['monition_mo']}",
        "Chargés d’annonce": f"FR — {payload['annonce_fr']}\nMO — {payload['annonce_mo']}",
    }


def generate_hybrid_month(state, year, month, services, manual_payloads, seed, unavailable_by_service=None, locks_by_service=None):
    """Génère un mois en respectant les deux modes de programmation.

    Les dimanches sont automatiques. Les célébrations supplémentaires utilisent
    les personnes choisies manuellement. Tout est traité chronologiquement afin
    que l'alternance individuelle reste exacte même lorsqu'une vigile se situe
    entre deux dimanches.
    """
    state = normalize_state(deepcopy(state))
    if history_has_month(state, year, month):
        raise RuntimeError("Ce mois est déjà validé dans l'historique.")
    rng = random.Random(seed + year * 100 + month)
    unavailable_by_service = unavailable_by_service or {}
    locks_by_service = locks_by_service or {}
    rows = []

    for service in sorted(services, key=lambda s: (s["date"], s.get("order", 20), s["id"])):
        today = date.fromisoformat(service["date"])
        if service.get("automatic_people"):
            unavailable = set(unavailable_by_service.get(service["id"], []))
            locks = locks_by_service.get(service["id"], {})
            f_read, m_read = choose_readers(state, today, unavailable, rng, locks)
            first = state["next_first_language"]
            if first == "FR":
                r1_code, r1_lang, r2_code, r2_lang = f_read, "FR", m_read, "MO"
            else:
                r1_code, r1_lang, r2_code, r2_lang = m_read, "MO", f_read, "FR"

            excluded = {f_read, m_read}
            f_mon, m_mon = choose_monitions(state, today, excluded, unavailable, rng, locks)
            excluded.update({f_mon, m_mon})
            f_ann = choose_announcement(state, "FR", today, excluded, unavailable, rng, locks.get("annonce_fr"))
            m_ann = choose_announcement(state, "MO", today, excluded, unavailable, rng, locks.get("annonce_mo"))

            assign(state, f_read, "LECTURE", today)
            assign(state, m_read, "LECTURE", today)
            state["reading_pairs"].append([f_read, m_read])
            assign(state, f_mon, "MONITION", today)
            assign(state, m_mon, "MONITION", today)
            state["monition_pairs"].append([f_mon, m_mon])
            assign(state, f_ann, "ANNONCE", today)
            assign(state, m_ann, "ANNONCE", today)
            state["next_first_language"] = "MO" if first == "FR" else "FR"

            refs = _form_standard_refs(service.get("form"))
            row = {
                "date": service["date"],
                "service_id": service["id"],
                "Dx": f"D{today.day}",
                "Réf.D": f"1re : {refs['r1']}\n2e : {refs['r2']}\nÉv. : {refs['ev']}",
                "Lecteurs": f"1re : {r1_lang} — {r1_code}\n2e : {r2_lang} — {r2_code}",
                "Monition introductive + P.U.": f"FR — {f_mon}\nMO — {m_mon}",
                "Chargés d’annonce": f"FR — {f_ann}\nMO — {m_ann}",
            }
        else:
            payload = manual_payloads.get(service["id"])
            if payload is None:
                raise RuntimeError(f"Affectations manuelles manquantes pour {service_display_title(service)}.")
            row = _apply_manual_service(state, service, payload)

        rows.append(row)
        state["history"].append(row)

    return rows, state

def history_has_month(state, year, month):
    prefix = f"{year:04d}-{month:02d}-"
    return any(str(row.get("date", "")).startswith(prefix) for row in state["history"])


def month_history(state, year, month):
    prefix = f"{year:04d}-{month:02d}-"
    return [row for row in state["history"] if str(row.get("date", "")).startswith(prefix)]


def reading_cycle_candidates(state, lang):
    codes = active_codes(state, lang)
    seen = set(state["reading_cycle_seen"][lang])
    if len(seen) == len(codes):
        state["reading_cycle_seen"][lang] = []
        seen = set()
    return [c for c in codes if c not in seen]


def reading_pool(state, lang, excluded, unavailable):
    candidates = reading_cycle_candidates(state, lang)
    return [
        c for c in candidates
        if c not in excluded
        and c not in unavailable
        and state["people"][c]["next_role"] in (None, "LECTURE")
    ]


def reading_rank(state, code, today):
    p = state["people"][code]
    return (
        p["reading_count"],
        -days_since(p["last_reading"], today),
        -days_since(p["last_service"], today),
        code,
    )


def choose_reader(state, lang, today, excluded, unavailable, rng, locked=None):
    pool = reading_pool(state, lang, excluded, unavailable)
    if locked:
        if locked not in active_codes(state, lang):
            raise RuntimeError(f"{display_name(state, locked)} n'appartient pas à la catégorie {lang}.")
        if locked in unavailable:
            raise RuntimeError(f"{display_name(state, locked)} est indisponible le {today.strftime('%d/%m/%Y')}.")
        if locked not in pool:
            nxt = ROLE_LABELS[state["people"][locked]["next_role"]]
            raise RuntimeError(
                f"Verrou impossible pour {display_name(state, locked)} en lecture le {today.strftime('%d/%m/%Y')}. "
                f"Sa prochaine fonction attendue est : {nxt}, ou son cycle de lecture n'est pas encore ouvert."
            )
        return locked
    if not pool:
        raise RuntimeError(
            f"Aucun lecteur {lang} disponible pour la lecture le {today.strftime('%d/%m/%Y')} "
            "sans casser la rotation individuelle ou le cycle d'équité."
        )
    rng.shuffle(pool)
    return min(pool, key=lambda c: reading_rank(state, c, today))


def monition_pool(state, lang, excluded, unavailable):
    return [
        c for c in active_codes(state, lang)
        if c not in excluded
        and c not in unavailable
        and state["people"][c]["next_role"] in (None, "MONITION")
    ]


def monition_rank(state, code, today):
    p = state["people"][code]
    return (
        p["monition_count"],
        -days_since(p["last_service"], today),
        code,
    )


def choose_monition(state, lang, today, excluded, unavailable, rng, locked=None):
    pool = monition_pool(state, lang, excluded, unavailable)
    if locked:
        if locked not in active_codes(state, lang):
            raise RuntimeError(f"{display_name(state, locked)} n'appartient pas à la catégorie {lang}.")
        if locked in unavailable:
            raise RuntimeError(f"{display_name(state, locked)} est indisponible le {today.strftime('%d/%m/%Y')}.")
        if locked not in pool:
            nxt = ROLE_LABELS[state["people"][locked]["next_role"]]
            raise RuntimeError(
                f"Verrou impossible pour {display_name(state, locked)} en monition/P.U. le {today.strftime('%d/%m/%Y')}. "
                f"Sa prochaine fonction attendue est : {nxt}."
            )
        return locked
    if not pool:
        raise RuntimeError(
            f"Aucun lecteur {lang} disponible pour monition/P.U. le {today.strftime('%d/%m/%Y')} "
            "sans casser l'alternance individuelle."
        )
    rng.shuffle(pool)
    return min(pool, key=lambda c: monition_rank(state, c, today))


def announcement_rank(state, code, today):
    p = state["people"][code]
    return (
        p["announcement_count"],
        -days_since(p["last_announcement"], today),
        code,
    )


def choose_announcement(state, lang, today, excluded, unavailable, rng, locked=None):
    pool = [
        c for c in active_codes(state, lang)
        if c not in excluded and c not in unavailable
    ]
    if locked:
        if locked not in active_codes(state, lang):
            raise RuntimeError(f"{display_name(state, locked)} n'appartient pas à la catégorie {lang}.")
        if locked in unavailable:
            raise RuntimeError(f"{display_name(state, locked)} est indisponible le {today.strftime('%d/%m/%Y')}.")
        if locked in excluded:
            raise RuntimeError(f"{display_name(state, locked)} a déjà une autre fonction le {today.strftime('%d/%m/%Y')}.")
        return locked
    if not pool:
        raise RuntimeError(f"Aucun {lang} disponible pour les annonces le {today.strftime('%d/%m/%Y')}.")
    rng.shuffle(pool)
    return min(pool, key=lambda c: announcement_rank(state, c, today))


def pair_penalty(pair, history_pairs):
    return 1 if list(pair) in history_pairs else 0


def choose_readers(state, today, unavailable, rng, locks):
    fr_locked = locks.get("lecture_fr")
    mo_locked = locks.get("lecture_mo")
    fr_pool = reading_pool(state, "FR", set(), unavailable)
    mo_pool = reading_pool(state, "MO", set(), unavailable)
    if fr_locked:
        fr_pool = [choose_reader(state, "FR", today, set(), unavailable, rng, fr_locked)]
    if mo_locked:
        mo_pool = [choose_reader(state, "MO", today, set(), unavailable, rng, mo_locked)]
    if not fr_pool or not mo_pool:
        raise RuntimeError(f"Impossible de constituer le binôme de lecture du {today.strftime('%d/%m/%Y')}.")

    choices = []
    for f in fr_pool:
        for m in mo_pool:
            score = (
                pair_penalty((f, m), state["reading_pairs"]),
                reading_rank(state, f, today),
                reading_rank(state, m, today),
                rng.random(),
            )
            choices.append((score, f, m))
    choices.sort(key=lambda x: x[0])
    return choices[0][1], choices[0][2]


def choose_monitions(state, today, excluded, unavailable, rng, locks):
    fr_locked = locks.get("monition_fr")
    mo_locked = locks.get("monition_mo")
    fr_pool = monition_pool(state, "FR", excluded, unavailable)
    mo_pool = monition_pool(state, "MO", excluded, unavailable)
    if fr_locked:
        fr_pool = [choose_monition(state, "FR", today, excluded, unavailable, rng, fr_locked)]
    if mo_locked:
        mo_pool = [choose_monition(state, "MO", today, excluded, unavailable, rng, mo_locked)]
    if not fr_pool or not mo_pool:
        raise RuntimeError(f"Impossible de constituer le binôme monition/P.U. du {today.strftime('%d/%m/%Y')}.")

    choices = []
    for f in fr_pool:
        for m in mo_pool:
            score = (
                pair_penalty((f, m), state["monition_pairs"]),
                monition_rank(state, f, today),
                monition_rank(state, m, today),
                rng.random(),
            )
            choices.append((score, f, m))
    choices.sort(key=lambda x: x[0])
    return choices[0][1], choices[0][2]


def assign(state, code, role, today):
    p = state["people"][code]
    if role == "LECTURE":
        p["reading_count"] += 1
        p["last_reading"] = today.isoformat()
        p["last_service"] = today.isoformat()
        p["next_role"] = "MONITION"
        lang = "FR" if code.startswith("F") else "MO"
        if code not in state["reading_cycle_seen"][lang]:
            state["reading_cycle_seen"][lang].append(code)
    elif role == "MONITION":
        p["monition_count"] += 1
        p["last_service"] = today.isoformat()
        p["next_role"] = "LECTURE"
    elif role == "ANNONCE":
        p["announcement_count"] += 1
        p["last_announcement"] = today.isoformat()


def parse_refs(text, dates):
    refs = {d.isoformat(): {"r1": "", "r2": "", "ev": ""} for d in dates}
    errors = []
    scheduled = set(refs)
    for line_no, raw in enumerate(text.splitlines(), start=1):
        raw = raw.strip()
        if not raw:
            continue
        parts = [p.strip() for p in raw.split("|")]
        if len(parts) < 4:
            errors.append(f"Ligne {line_no}: format incomplet")
            continue
        try:
            date.fromisoformat(parts[0])
        except ValueError:
            errors.append(f"Ligne {line_no}: date invalide ({parts[0]})")
            continue
        if parts[0] in scheduled:
            refs[parts[0]] = {"r1": parts[1], "r2": parts[2], "ev": parts[3]}

    for d in dates:
        item = refs[d.isoformat()]
        missing = [label for label, key in (("1re", "r1"), ("2e", "r2"), ("Évangile", "ev")) if not item[key]]
        if missing:
            errors.append(f"{d.strftime('%d/%m/%Y')}: référence(s) manquante(s) — {', '.join(missing)}")
    return refs, errors


def generate_month(state, year, month, refs, seed, unavailable_by_date=None, locks_by_date=None, service_dates=None, celebrations=None):
    state = normalize_state(deepcopy(state))
    rng = random.Random(seed + year * 100 + month)
    unavailable_by_date = unavailable_by_date or {}
    locks_by_date = locks_by_date or {}

    if history_has_month(state, year, month):
        raise RuntimeError(
            "Ce mois est déjà validé dans l'historique. Importez un état antérieur ou réinitialisez la rotation."
        )

    rows = []
    service_dates = sorted(service_dates or sundays(year, month))
    celebrations = celebrations or {}
    for service_day in service_dates:
        key = service_day.isoformat()
        unavailable = set(unavailable_by_date.get(key, []))
        locks = locks_by_date.get(key, {})

        f_read, m_read = choose_readers(state, service_day, unavailable, rng, locks)
        first = state["next_first_language"]
        if first == "FR":
            r1_code, r1_lang, r2_code, r2_lang = f_read, "FR", m_read, "MO"
        else:
            r1_code, r1_lang, r2_code, r2_lang = m_read, "MO", f_read, "FR"

        excluded = {f_read, m_read}
        f_mon, m_mon = choose_monitions(state, service_day, excluded, unavailable, rng, locks)
        excluded.update({f_mon, m_mon})

        f_ann = choose_announcement(
            state, "FR", service_day, excluded, unavailable, rng, locks.get("annonce_fr")
        )
        m_ann = choose_announcement(
            state, "MO", service_day, excluded, unavailable, rng, locks.get("annonce_mo")
        )

        assign(state, f_read, "LECTURE", service_day)
        assign(state, m_read, "LECTURE", service_day)
        state["reading_pairs"].append([f_read, m_read])
        assign(state, f_mon, "MONITION", service_day)
        assign(state, m_mon, "MONITION", service_day)
        state["monition_pairs"].append([f_mon, m_mon])
        assign(state, f_ann, "ANNONCE", service_day)
        assign(state, m_ann, "ANNONCE", service_day)
        state["next_first_language"] = "MO" if first == "FR" else "FR"

        ref = refs.get(key, {"r1": "", "r2": "", "ev": ""})
        celebration = _compact_spaces(celebrations.get(key, ""))
        day_code = f"D{service_day.day}" if service_day.weekday() == 6 else f"J{service_day.day}"
        ref_lines = []
        if celebration:
            ref_lines.append(f"⭐ {celebration}" if service_day.weekday() != 6 else celebration)
        ref_lines.extend([f"1re : {ref['r1']}", f"2e : {ref['r2']}", f"Év. : {ref['ev']}"])
        row = {
            "date": key,
            "Dx": day_code,
            "Réf.D": "\n".join(ref_lines),
            "Lecteurs": f"1re : {r1_lang} — {r1_code}\n2e : {r2_lang} — {r2_code}",
            "Monition introductive + P.U.": f"FR — {f_mon}\nMO — {m_mon}",
            "Chargés d’annonce": f"FR — {f_ann}\nMO — {m_ann}",
        }
        rows.append(row)
        state["history"].append(row)

    return rows, state


# -----------------------------------------------------------------------------
# Exports
# -----------------------------------------------------------------------------
def rows_for_export(rows, state):
    fields = ["Dx", "Réf.D", "Lecteurs", "Monition introductive + P.U.", "Chargés d’annonce"]
    shown_rows = [display_row(row, state) for row in rows]
    return fields, [{k: row.get(k, "") for k in fields} for row in shown_rows]


def excel_data(rows, state, title="Programme liturgique"):
    fields, clean_rows = rows_for_export(rows, state)
    wb = Workbook()
    ws = wb.active
    ws.title = "Programme"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(fields))
    ws.cell(1, 1, title)
    ws.cell(1, 1).font = Font(bold=True, size=16)
    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    for col, field in enumerate(fields, start=1):
        cell = ws.cell(3, col, field)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_idx, row in enumerate(clean_rows, start=4):
        for c_idx, field in enumerate(fields, start=1):
            cell = ws.cell(r_idx, c_idx, row[field])
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r_idx].height = 58

    widths = [9, 30, 28, 30, 24]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A4"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def pdf_data(rows, state, title="Programme liturgique"):
    out = io.BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=landscape(A4),
        rightMargin=8 * mm,
        leftMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    styles = getSampleStyleSheet()
    normal = styles["BodyText"]
    normal.fontName = "Helvetica"
    normal.fontSize = 7.5
    normal.leading = 9
    head = styles["Heading2"]
    head.alignment = 1

    fields, clean_rows = rows_for_export(rows, state)
    data = [[Paragraph(f"<b>{html.escape(f)}</b>", normal) for f in fields]]
    for row in clean_rows:
        data.append([
            Paragraph(html.escape(str(row[f])).replace("\n", "<br/>"), normal)
            for f in fields
        ])

    table = Table(data, colWidths=[18 * mm, 56 * mm, 50 * mm, 56 * mm, 48 * mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.35, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAD3")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    doc.build([Paragraph(title, head), Spacer(1, 4 * mm), table])
    return out.getvalue()


def json_state_bytes(state):
    return json.dumps(state, ensure_ascii=False, indent=2).encode("utf-8")


# -----------------------------------------------------------------------------
# Rendu mobile
# -----------------------------------------------------------------------------
def render_mobile_cards(rows, state):
    for raw_row in rows:
        row = display_row(raw_row, state)
        st.markdown(
            f"""
            <div class="lit-card">
              <div class="lit-date">{html.escape(row['Dx'])}</div>
              <div class="lit-grid">
                <div class="lit-section"><div class="lit-label">Réf.D</div>{html.escape(row['Réf.D']).replace(chr(10), '<br>')}</div>
                <div class="lit-section"><div class="lit-label">Lecteurs</div>{html.escape(row['Lecteurs']).replace(chr(10), '<br>')}</div>
                <div class="lit-section"><div class="lit-label">Monition + P.U.</div>{html.escape(row['Monition introductive + P.U.']).replace(chr(10), '<br>')}</div>
                <div class="lit-section"><div class="lit-label">Annonces</div>{html.escape(row['Chargés d’annonce']).replace(chr(10), '<br>')}</div>
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )


def availability_key(year, month, day):
    return f"availability_{year}_{month}_{day}"


def lock_key(year, month, day, role):
    return f"lock_{year}_{month}_{day}_{role}"


def ensure_session():
    if "scheduler_state" not in st.session_state:
        st.session_state.scheduler_state = initial_state()
    if "draft_rows" not in st.session_state:
        st.session_state.draft_rows = []
    if "draft_state" not in st.session_state:
        st.session_state.draft_state = None
    if "draft_meta" not in st.session_state:
        st.session_state.draft_meta = None
    if "generation_nonce" not in st.session_state:
        st.session_state.generation_nonce = 0


ensure_session()

# -----------------------------------------------------------------------------
# Interface
# -----------------------------------------------------------------------------
def clear_draft():
    st.session_state.draft_rows = []
    st.session_state.draft_state = None
    st.session_state.draft_meta = None


st.title("⛪ Programme liturgique")
st.caption(f"Version : {APP_VERSION}")
st.caption("Programmation automatique — les codes techniques restent en arrière-plan, seuls les noms sont affichés")

home, generator, members_tab, history_tab, rotation_tab, settings_tab = st.tabs([
    "🏠 Accueil", "✨ Générer", "👥 Membres", "🕘 Historique", "🔄 Rotation", "⚙️ Réglages"
])

# -----------------------------------------------------------------------------
# ACCUEIL
# -----------------------------------------------------------------------------
with home:
    state = st.session_state.scheduler_state
    c1, c2, c3 = st.columns(3)
    c1.metric("Francophones", len(active_codes(state, "FR")))
    c2.metric("Mooréphones", len(active_codes(state, "MO")))
    c3.metric("Célébrations validées", len(state["history"]))

    missing = missing_active_names(state)
    if missing:
        st.warning(
            f"Il reste {len(missing)} nom(s) à renseigner. Ouvrez l’onglet 👥 Membres avant de générer un programme."
        )
    else:
        st.success("Tous les membres actifs ont un nom. L’application est prête à générer un programme.")

    st.markdown("### Règles actives")
    st.markdown(
        "- **Dimanches :** affectations automatiques et alternance FR ↔ MO des 1re/2e lectures.\n"
        "- **Par personne :** Lecture ↔ Monition/P.U. à son prochain passage, y compris après une fête ou une vigile.\n"
        "- **Lectures :** priorité à l’équité du cycle et à l’espacement des passages.\n"
        "- **Binômes :** éviter de reformer les mêmes couples FR–MO.\n"
        "- **Annonces :** indépendantes de l’alternance Lecture/Monition.\n"
        "- **Même célébration :** pas de cumul de deux fonctions.\n"
        "- **Références bibliques :** automatiques via l’AELF pour dimanches, grandes fêtes et vigiles ; noms manuels hors dimanche."
    )

    if st.session_state.draft_rows:
        st.info("Un brouillon est en attente de validation dans l’onglet ✨ Générer.")

    st.markdown("### Sauvegarde recommandée")
    st.caption(
        "Après avoir saisi les noms ou validé un mois, téléchargez une sauvegarde JSON dans ⚙️ Réglages. "
        "Elle permet de restaurer les noms, l’historique et la rotation après un redémarrage de l’application."
    )

# -----------------------------------------------------------------------------
# GENERATEUR
# -----------------------------------------------------------------------------
with generator:
    state = st.session_state.scheduler_state
    names_ready = not missing_active_names(state)

    if not names_ready:
        st.error("Renseignez d’abord tous les noms des membres actifs dans 👥 Membres.")

    st.markdown("### 1. Choisir la période")
    c1, c2 = st.columns(2)
    with c1:
        year = int(st.number_input("Année", 2020, 2100, 2026, 1, key="gen_year"))
    with c2:
        month = st.selectbox(
            "Mois", range(1, 13), index=8,
            format_func=lambda m: MONTHS[m - 1], key="gen_month"
        )

    already_validated = history_has_month(state, year, month)
    if already_validated:
        st.warning("Ce mois est déjà validé dans l’historique. Vous pouvez le consulter dans 🕘 Historique.")

    st.markdown("### 2. Calendrier et références bibliques automatiques")
    st.caption(
        "Les dimanches sont programmés automatiquement. Les grandes célébrations en semaine et leurs vigiles "
        "sont signalées avec leurs références AELF, mais leurs personnes sont choisies manuellement. "
        "La Veillée pascale et les messes de la veille / de la nuit de Noël sont prises en compte."
    )

    services_key = f"hybrid_services_{year}_{month}"
    reload_calendar = st.button("🔄 Recharger le calendrier depuis l’AELF", key=f"reload_hybrid_{year}_{month}")
    if services_key not in st.session_state or reload_calendar:
        with st.spinner("Analyse du calendrier liturgique, des grandes fêtes et des vigiles…"):
            st.session_state[services_key] = discover_hybrid_services(year, month)

    services = st.session_state.get(services_key, [])
    sunday_services = [s for s in services if s.get("automatic_people")]
    extra_services = [s for s in services if not s.get("automatic_people")]

    st.caption(
        f"Détecté : {len(sunday_services)} dimanche(s) automatique(s) et "
        f"{len(extra_services)} célébration(s) supplémentaire(s) à affectation manuelle."
    )

    if not sunday_services:
        st.error("Aucun dimanche n’a pu être chargé depuis l’AELF. Appuyez sur Recharger.")

    with st.expander("⛪ Messes dominicales — programmation automatique", expanded=True):
        for service in sunday_services:
            d = date.fromisoformat(service["date"])
            status = "✅" if service.get("status") == "ok" else "⚠️"
            st.markdown(f"{status} **D{d.day} — {service_display_title(service)}**")
            st.caption(service_reference_text(service) or service.get("error", "Références indisponibles"))

    st.markdown("### 3. Journées supplémentaires — noms manuels")
    st.info(
        "Ces célébrations n’avancent pas l’alternance FR/MO des dimanches. En revanche, un passage manuel en Lecture "
        "ou en Monition/P.U. compte bien pour l’alternance individuelle de la personne."
    )

    included_extra_ids = []
    manual_payloads = {}
    if extra_services:
        for service in extra_services:
            d = date.fromisoformat(service["date"])
            default = service.get("default_include", True)
            include = st.checkbox(
                f"{d.strftime('%d/%m/%Y')} — {service_display_title(service)}",
                value=default,
                key=_service_key(service, "include_service"),
            )
            if not include:
                continue
            included_extra_ids.append(service["id"])
            with st.expander(f"✍️ Affecter les personnes — {d.strftime('%d/%m')} — {service.get('form_title', '')}", expanded=False):
                if names_ready:
                    manual_payloads[service["id"]] = collect_manual_service_payload(service, state)
                else:
                    st.warning("Saisissez d’abord les noms dans 👥 Membres.")
    else:
        st.caption("Aucune célébration supplémentaire ni vigile détectée pour ce mois.")

    selected_services = sunday_services + [s for s in extra_services if s["id"] in included_extra_ids]
    selected_services.sort(key=lambda s: (s["date"], s.get("order", 20), s["id"]))

    with st.expander("🚫 Indisponibilités des dimanches", expanded=False):
        if names_ready:
            all_active = active_codes(state, "FR") + active_codes(state, "MO")
            for service in sunday_services:
                d = date.fromisoformat(service["date"])
                st.multiselect(
                    f"D{d.day} — {d.strftime('%d/%m/%Y')}", all_active,
                    key=f"sun_unavailable_{service['id']}", placeholder="Aucune indisponibilité",
                    format_func=member_format(state),
                )
        else:
            st.info("Disponible après la saisie des noms.")

    with st.expander("🔒 Verrouiller une affectation dominicale", expanded=False):
        st.caption("Optionnel : laissez Auto pour que l’algorithme choisisse.")
        if names_ready:
            for service in sunday_services:
                d = date.fromisoformat(service["date"])
                st.markdown(f"**D{d.day} — {d.strftime('%d/%m/%Y')}**")
                a, b = st.columns(2)
                with a:
                    st.selectbox("Lecture FR", ["Auto"] + active_codes(state, "FR"), key=f"sun_lock_{service['id']}_lecture_fr", format_func=member_format(state))
                    st.selectbox("Monition FR", ["Auto"] + active_codes(state, "FR"), key=f"sun_lock_{service['id']}_monition_fr", format_func=member_format(state))
                    st.selectbox("Annonce FR", ["Auto"] + active_codes(state, "FR"), key=f"sun_lock_{service['id']}_annonce_fr", format_func=member_format(state))
                with b:
                    st.selectbox("Lecture MO", ["Auto"] + active_codes(state, "MO"), key=f"sun_lock_{service['id']}_lecture_mo", format_func=member_format(state))
                    st.selectbox("Monition MO", ["Auto"] + active_codes(state, "MO"), key=f"sun_lock_{service['id']}_monition_mo", format_func=member_format(state))
                    st.selectbox("Annonce MO", ["Auto"] + active_codes(state, "MO"), key=f"sun_lock_{service['id']}_annonce_mo", format_func=member_format(state))
                st.divider()

    st.markdown("### 4. Générer")
    seed = int(st.number_input("Graine de brassage", 0, 999999, 2026, 1, key="seed"))
    st.caption(f"Prochaine 1re lecture dominicale prévue : **{state['next_first_language']}**")

    unavailable_by_service = {}
    locks_by_service = {}
    for service in sunday_services:
        unavailable_by_service[service["id"]] = st.session_state.get(f"sun_unavailable_{service['id']}", []) if names_ready else []
        day_locks = {}
        if names_ready:
            for role in ["lecture_fr", "lecture_mo", "monition_fr", "monition_mo", "annonce_fr", "annonce_mo"]:
                value = st.session_state.get(f"sun_lock_{service['id']}_{role}", "Auto")
                if value != "Auto":
                    day_locks[role] = value
        locks_by_service[service["id"]] = day_locks

    generate_clicked = st.button(
        "✨ Générer un brouillon",
        type="primary",
        disabled=(not names_ready or already_validated or not sunday_services),
    )
    if generate_clicked:
        try:
            effective_seed = seed + st.session_state.generation_nonce
            rows, draft_state = generate_hybrid_month(
                state, year, month, selected_services, manual_payloads, effective_seed,
                unavailable_by_service, locks_by_service,
            )
            st.session_state.draft_rows = rows
            st.session_state.draft_state = draft_state
            st.session_state.draft_meta = {"year": year, "month": month, "seed": effective_seed}
            st.success("Brouillon généré. Vérifiez-le avant validation.")
        except Exception as exc:
            st.error(str(exc))

    if st.session_state.draft_rows:
        meta = st.session_state.draft_meta or {}
        st.markdown("### 5. Brouillon")
        st.caption(
            f"{MONTHS[meta.get('month', month)-1]} {meta.get('year', year)} — non encore comptabilisé dans la rotation"
        )
        render_mobile_cards(st.session_state.draft_rows, st.session_state.draft_state or state)

        c1, c2 = st.columns(2)
        with c1:
            if st.button("🔀 Régénérer"):
                st.session_state.generation_nonce += 1
                clear_draft()
                st.info("La graine de brassage a été modifiée. Appuyez sur Générer un brouillon.")
        with c2:
            if st.button("✅ Valider ce mois", type="primary"):
                st.session_state.scheduler_state = st.session_state.draft_state
                clear_draft()
                st.success("Programme validé et ajouté à l’historique.")
                st.rerun()

        st.markdown("### 6. Exporter le brouillon")
        draft_state = st.session_state.draft_state or state
        ex1, ex2 = st.columns(2)
        with ex1:
            st.download_button(
                "📊 Excel",
                excel_data(st.session_state.draft_rows, draft_state, "Programme liturgique — Brouillon"),
                "programme_liturgique_brouillon.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with ex2:
            st.download_button(
                "📄 PDF",
                pdf_data(st.session_state.draft_rows, draft_state, "Programme liturgique — Brouillon"),
                "programme_liturgique_brouillon.pdf",
                "application/pdf",
            )

# -----------------------------------------------------------------------------
# MEMBRES
# -----------------------------------------------------------------------------
with members_tab:
    state = st.session_state.scheduler_state
    st.markdown("### Gestion des membres")
    st.caption(
        "Vous travaillez uniquement avec les noms. Un identifiant interne est créé automatiquement et reste invisible. "
        "Il garantit que l’historique d’une personne reste correct même si son nom est corrigé plus tard."
    )

    c1, c2 = st.columns(2)
    c1.metric("Francophones actifs", len(active_codes(state, "FR")))
    c2.metric("Mooréphones actifs", len(active_codes(state, "MO")))

    missing = missing_active_names(state)
    if missing:
        st.markdown("#### ✍️ Première saisie des noms")
        st.info(
            "Saisissez un nom par ligne. Cette étape associe les noms aux identifiants internes déjà créés. "
            "Après l’initialisation, les identifiants ne seront plus nécessaires à l’écran."
        )
        b1, b2 = st.columns(2)
        with b1:
            fr_text = st.text_area(
                f"Francophones — {len(active_codes(state, 'FR'))} noms",
                value="\n".join(member_name(state, c) for c in active_codes(state, "FR")),
                height=300,
                placeholder="Un nom par ligne",
                key="bulk_names_fr",
            )
        with b2:
            mo_text = st.text_area(
                f"Mooréphones — {len(active_codes(state, 'MO'))} noms",
                value="\n".join(member_name(state, c) for c in active_codes(state, "MO")),
                height=300,
                placeholder="Un nom par ligne",
                key="bulk_names_mo",
            )
        if st.button("💾 Enregistrer tous les noms", type="primary"):
            try:
                fr_names = fr_text.splitlines()
                mo_names = mo_text.splitlines()
                apply_bulk_names(state, fr_names, mo_names)
                clear_draft()
                st.success("Les noms ont été enregistrés.")
                st.rerun()
            except Exception as exc:
                st.error(str(exc))
    else:
        active_rows = []
        for code in active_codes(state, "FR"):
            active_rows.append({"Nom": display_name(state, code), "Catégorie": "Francophone", "Statut": "Actif"})
        for code in active_codes(state, "MO"):
            active_rows.append({"Nom": display_name(state, code), "Catégorie": "Mooréphone", "Statut": "Actif"})
        st.dataframe(active_rows, use_container_width=True, hide_index=True)

        st.markdown("#### ➕ Ajouter un membre")
        add_cat = st.radio("Catégorie du nouveau membre", ["Francophone", "Mooréphone"], horizontal=True, key="add_cat")
        new_name = st.text_input("Nom du nouveau membre", placeholder="Ex. Mme Dupont", key="new_member_name")
        if st.button("➕ Ajouter le membre", type="primary"):
            try:
                lang = "FR" if add_cat == "Francophone" else "MO"
                add_member(state, lang, new_name)
                clear_draft()
                st.success(f"{clean_name(new_name)} a été ajouté(e).")
                st.rerun()
            except Exception as exc:
                st.error(str(exc))

        st.markdown("#### ✏️ Corriger un nom")
        all_known = active_codes(state, "FR") + active_codes(state, "MO") + state.get("inactive_members", {}).get("FR", []) + state.get("inactive_members", {}).get("MO", [])
        all_known = sort_codes(all_known)
        edit_code = st.selectbox("Membre", all_known, format_func=member_format(state), key="edit_member_select")
        corrected_name = st.text_input(
            "Nouveau nom",
            value=member_name(state, edit_code),
            key=f"edit_member_name_{edit_code}",
        )
        if st.button("💾 Enregistrer la correction"):
            try:
                old_name = display_name(state, edit_code)
                set_member_name(state, edit_code, corrected_name)
                clear_draft()
                st.success(f"Nom corrigé : {old_name} → {clean_name(corrected_name)}.")
                st.rerun()
            except Exception as exc:
                st.error(str(exc))

        st.markdown("#### ➖ Retirer temporairement un membre")
        st.caption("Le membre n’est plus programmé, mais son historique et sa prochaine fonction sont conservés.")
        d1, d2 = st.columns(2)
        with d1:
            fr_active = active_codes(state, "FR")
            fr_remove = st.selectbox(
                "Francophone à retirer", fr_active, format_func=member_format(state), key="remove_fr"
            )
            if st.button("Retirer ce francophone", disabled=len(fr_active) <= 3):
                try:
                    name = display_name(state, fr_remove)
                    remove_member(state, fr_remove)
                    clear_draft()
                    st.success(f"{name} est maintenant inactif(ve). Son historique est conservé.")
                    st.rerun()
                except Exception as exc:
                    st.error(str(exc))
        with d2:
            mo_active = active_codes(state, "MO")
            mo_remove = st.selectbox(
                "Mooréphone à retirer", mo_active, format_func=member_format(state), key="remove_mo"
            )
            if st.button("Retirer ce mooréphone", disabled=len(mo_active) <= 3):
                try:
                    name = display_name(state, mo_remove)
                    remove_member(state, mo_remove)
                    clear_draft()
                    st.success(f"{name} est maintenant inactif(ve). Son historique est conservé.")
                    st.rerun()
                except Exception as exc:
                    st.error(str(exc))

        inactive_codes = state.get("inactive_members", {}).get("FR", []) + state.get("inactive_members", {}).get("MO", [])
        if inactive_codes:
            st.markdown("#### ♻️ Réactiver un membre")
            react_code = st.selectbox(
                "Membre inactif", inactive_codes, format_func=member_format(state), key="reactivate_member"
            )
            if st.button("♻️ Réactiver"):
                try:
                    name = display_name(state, react_code)
                    reactivate_member(state, react_code)
                    clear_draft()
                    st.success(f"{name} est de nouveau actif(ve).")
                    st.rerun()
                except Exception as exc:
                    st.error(str(exc))

        inactive_rows = []
        for code in state.get("inactive_members", {}).get("FR", []):
            inactive_rows.append({"Nom": display_name(state, code), "Catégorie": "Francophone", "Statut": "Inactif"})
        for code in state.get("inactive_members", {}).get("MO", []):
            inactive_rows.append({"Nom": display_name(state, code), "Catégorie": "Mooréphone", "Statut": "Inactif"})
        if inactive_rows:
            with st.expander("Voir les membres inactifs"):
                st.dataframe(inactive_rows, use_container_width=True, hide_index=True)

# -----------------------------------------------------------------------------
# HISTORIQUE
# -----------------------------------------------------------------------------
with history_tab:
    state = st.session_state.scheduler_state
    st.markdown("### Historique validé")
    hist = state["history"]
    if not hist:
        st.info("Aucun mois n’est encore validé.")
    else:
        years = sorted({int(r["date"][:4]) for r in hist})
        hy = st.selectbox("Année", years, index=len(years) - 1, key="hist_year")
        months_available = sorted({int(r["date"][5:7]) for r in hist if int(r["date"][:4]) == hy})
        hm = st.selectbox(
            "Mois", months_available, index=len(months_available) - 1,
            format_func=lambda m: MONTHS[m - 1], key="hist_month"
        )
        rows = month_history(state, hy, hm)
        render_mobile_cards(rows, state)
        st.markdown("#### Export")
        a, b = st.columns(2)
        with a:
            st.download_button(
                "📊 Excel",
                excel_data(rows, state, f"Programme liturgique — {MONTHS[hm-1]} {hy}"),
                f"programme_{hy}_{hm:02d}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with b:
            st.download_button(
                "📄 PDF",
                pdf_data(rows, state, f"Programme liturgique — {MONTHS[hm-1]} {hy}"),
                f"programme_{hy}_{hm:02d}.pdf",
                "application/pdf",
            )

# -----------------------------------------------------------------------------
# ROTATION
# -----------------------------------------------------------------------------
with rotation_tab:
    state = st.session_state.scheduler_state
    st.markdown("### Contrôle individuel")
    if missing_active_names(state):
        st.info("Le tableau sera plus lisible après la saisie de tous les noms.")
    group_filter = st.radio("Catégorie", ["Tous", "FR", "MO"], horizontal=True)
    codes = (
        active_codes(state, "FR") + active_codes(state, "MO")
        if group_filter == "Tous" else active_codes(state, group_filter)
    )
    control = []
    for code in codes:
        p = state["people"][code]
        control.append({
            "Nom": display_name(state, code),
            "Prochaine fonction": ROLE_LABELS[p["next_role"]],
            "Lectures": p["reading_count"],
            "Monitions/P.U.": p["monition_count"],
            "Annonces": p["announcement_count"],
            "Dernière lecture": p["last_reading"] or "—",
            "Dernier passage L/M": p["last_service"] or "—",
        })
    st.dataframe(control, use_container_width=True, hide_index=True)

    st.markdown("### Cycle de lecture en cours")
    fr_seen = [display_name(state, c) for c in state["reading_cycle_seen"]["FR"]]
    mo_seen = [display_name(state, c) for c in state["reading_cycle_seen"]["MO"]]
    st.write("FR déjà passés :", ", ".join(fr_seen) or "—")
    st.write("MO déjà passés :", ", ".join(mo_seen) or "—")

# -----------------------------------------------------------------------------
# REGLAGES / SAUVEGARDE
# -----------------------------------------------------------------------------
with settings_tab:
    state = st.session_state.scheduler_state
    st.markdown("### Sauvegarder / restaurer")
    st.caption("La sauvegarde contient les noms, les membres actifs/inactifs, l’historique et l’état de la rotation.")
    st.download_button(
        "💾 Télécharger la sauvegarde (JSON)",
        json_state_bytes(state),
        "etat_programmation.json",
        "application/json",
    )

    uploaded = st.file_uploader("Restaurer une sauvegarde JSON", type=["json"], key="restore_json")
    if uploaded is not None:
        try:
            restored = normalize_state(json.load(uploaded))
            st.session_state.scheduler_state = restored
            clear_draft()
            st.success("Sauvegarde restaurée.")
            st.rerun()
        except Exception as exc:
            st.error(f"Import impossible : {exc}")

    st.markdown("### Première langue")
    current = state["next_first_language"]
    chosen = st.radio(
        "Langue de la prochaine 1re lecture",
        ["FR", "MO"], horizontal=True,
        index=0 if current == "FR" else 1,
        key="first_language_setting",
    )
    if chosen != current:
        state["next_first_language"] = chosen
        clear_draft()
        st.success(f"Prochaine 1re lecture réglée sur {chosen}.")

    st.markdown("### Réinitialiser uniquement la rotation")
    st.caption("Conserve les membres et leurs noms, mais efface l’historique, les compteurs et les derniers passages.")
    confirm_rotation_reset = st.checkbox("Je confirme la remise à zéro de la rotation", key="confirm_rotation_reset")
    if st.button("🔄 Réinitialiser la rotation", disabled=not confirm_rotation_reset):
        st.session_state.scheduler_state = reset_rotation_keep_members(state)
        clear_draft()
        st.success("Rotation réinitialisée. Les membres et les noms sont conservés.")
        st.rerun()

    st.markdown("### Réinitialisation complète")
    st.warning("Cette opération efface aussi tous les noms et revient au groupe initial de 10 francophones et 8 mooréphones.")
    confirm_reset = st.checkbox("Je confirme vouloir tout effacer", key="confirm_full_reset")
    if st.button("🗑️ Tout réinitialiser", disabled=not confirm_reset):
        st.session_state.scheduler_state = initial_state()
        clear_draft()
        st.success("Application réinitialisée.")
        st.rerun()

st.markdown("---")
st.markdown(
    '<div class="small-note">Conseil mobile : ajoutez la page à votre écran d’accueil. Pensez aussi à télécharger régulièrement la sauvegarde JSON.</div>',
    unsafe_allow_html=True,
)
